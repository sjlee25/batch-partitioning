import tvm
from tvm import relay
from tvm.contrib import graph_runtime
from util import get_network
import argparse
import copy
import numpy as np
import openpyxl
from os import path,_exit
import sys
import time
import threading
from Inference import Device, Environment
from Partition import Partitioner

class TVMGraph:
    def __init__(self, env, dev):
        self.dev = dev
        net, params, input_shape, output_shape = \
            get_network(name=env.network, batch_size=dev.batch_size)
        with relay.build_config(opt_level=env.opt_level):
            graph, lib, params = relay.build(net, target=dev.target, params=params)
        data = tvm.nd.array((np.random.uniform(size=input_shape)).astype('float32'))

        self.module = graph_runtime.create(graph, lib, dev.ctx)
        self.module.set_input('data', data, **params)

    def Run(self):
        timer = self.module.module.time_evaluator('run', self.dev.ctx, number=1, repeat=1)
        self.dev.eval_time = np.mean(np.array(timer().results)) * 1000

class ParamTester:
    def __init__(self, env, interval=0.05, log_path=''):
        self.env = env
        self.denoise_thresh = 1.5
        self.offload_thresh = 1.5
        self.devices = env.devices
        self.network = env.network
        self.batch_size = env.batch_size
        self.repeat = 3
        self.interval = interval
        self.log_path = log_path

        self.partitioner = Partitioner(env, self.denoise_thresh, self.offload_thresh)

    def DenoiseTest(self):
        pass

    def OffloadTest(self):
        current_thresh = self.offload_thresh
        min_time = float('inf')
        std = 0.0
            
        while current_thresh >= 1:
            threads = []
            dev_times = []
            self.partitioner.offload_thresh = current_thresh
            self.partitioner.StartPartition()
            
            for dev in self.devices:
                if dev.batch_size > 0:
                    t = threading.Thread(target=TVMGraph(self.env, dev).Run)
                    threads.append(t)
                    t.start()
            for t in threads:
                t.join()

            for dev in self.devices:
                dev_times.append(dev.eval_time)
            cur_time = max(dev_times)
            
            if cur_time <= min_time:
                self.offload_thresh = current_thresh
                min_time = cur_time
                std = np.array(dev_times).std()
            current_thresh -= self.interval

        print(self.network, self.batch_size, self.offload_thresh, min_time)

        if env.log_path != '':
            sheet_name = 'offload'
            if path.exists(self.log_path):
                book = openpyxl.load_workbook(self.log_path)
                if sheet_name in book:
                    sheet = book[sheet_name]
                else: sheet = book.create_sheet(sheet_name)
            else:
                book = openpyxl.Workbook()
                sheet = book.create_sheet(sheet_name)
            max_row = sheet.max_row
            sheet.cell(row=max_row+1, column=1, value=network)
            sheet.cell(row=max_row+1, column=2, value=batch_size)
            sheet.cell(row=max_row+1, column=3, value=self.offload_thresh)
            sheet.cell(row=max_row+1, column=4, value=min_time)
            sheet.cell(row=max_row+1, column=5, value=std)
            book.save(self.log_path)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--network', type=str, choices=
                        ['mobilenet', 'resnet-18', 'resnet-34', 'resnet-50',
                        'vgg-16', 'vgg-19', 'inception_v3', 'densenet-121',
                        'squeezenet_v1.0', 'squeezenet_v1.1'], 
                        help='The name of neural network to inference')
    parser.add_argument('--batch', type=int, help='Max Batch size')
    parser.add_argument('--device', type=str, default='gpu0',
                        help='Devices to use, give as \'cpu\', \'igpu\' or \'gpu0\' successively')
    parser.add_argument('--log', type=str, default='',
                        help='File path for logging')
    args = parser.parse_args()

    network = args.network
    batch_size = args.batch
    arg_devs = list(args.device.split(','))

    devices = []
    gpus = []
    cpu = igpu = None
    cpu_idx = igpu_idx = -1

    # Check available devices
    for dev in arg_devs:
        if dev == 'cpu':
            if tvm.cpu(0).exist:
                cpu = Device('cpu', 0)
            else: print('[Error] Device \'%s\' is unrecognizable' % dev)

        elif dev == 'igpu':
            if tvm.opencl(0).exist:
                igpu = Device('igpu', 0)
            else: print('[Error] Device \'%s\' is unrecognizable' % dev)

        elif dev.find('gpu') >= 0:
            idx_start = dev.find('gpu') + len('gpu')
            gpu_idx = int(dev[idx_start:])
            if tvm.gpu(gpu_idx).exist:
                gpus.append(Device('gpu', gpu_idx))
            else: print('[Error] Device \'%s\' is unrecognizable' % dev)

    # Add devices
    if cpu is not None:
        cpu_idx = len(devices)
        devices.append(cpu)
    if igpu is not None:
        igpu_idx = len(devices)
        devices.append(igpu)
    if len(gpus) > 0:
        devices.extend(gpus)

    if batch_size == 0 or len(devices) == 0:
        parser.print_help(sys.stderr)
        exit(1)

    env = Environment(network, batch_size, devices, args.log)
    tester = ParamTester(env, 0.05, args.log)
    tester.OffloadTest()
