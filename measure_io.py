import tvm
from tvm import relay, rpc
from tvm.contrib import graph_runtime, util
from util import get_network
import argparse
import numpy as np
import sys
import time
import openpyxl
from os import path, _exit
import threading
# from Inference import Environment, Device
from Inference import Environment

def connectRPC(ctx, remote, graph, lib, params, data):
    global input_time, output_time

    temp = util.tempdir()
    path = temp.relpath('lib.tar')
    lib.export_library(path)
    remote.upload(path)

    LIB = remote.load_module('lib.tar')
    module = graph_runtime.create(graph, LIB, ctx)
    # module.set_input(**params)

    input_time = time.time()
    module.set_input('data', data)
    ctx.sync()
    input_time = (time.time() - input_time) * 1000 # ms

    output_time = time.time()
    module.get_output(0)
    ctx.sync()
    output_time = (time.time() - output_time) * 1000 # ms

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--network', type=str, choices=
                        ['mobilenet', 'squeezenet_v1.0', 'squeezenet_v1.1', 'resnet-18', 'resnet-34',
                        'resnet-50', 'inception_v3', 'vgg-16', 'vgg-19', 'densenet-121'], 
                        help='Name of neural network model to use')
    parser.add_argument('--device', type=str, default='gpu0',
                        help='Device to use, write just one argument as \'cpu\', \'igpu\' or \'gpu0\'')
    parser.add_argument('--batch', type=int, help='Batch size')
    parser.add_argument('--max_batch', type=int, default=0, help='Max batch size for iterating check')
    parser.add_argument('--inc', type=int, default=0, help='Increment of batch size for iterating check')
    parser.add_argument('--log', type=str, default='', help='File path for logging')
    args = parser.parse_args()

    batch_size = args.batch
    max_batch = args.max_batch
    increment = args.inc

    if batch_size == 0 or increment < 0:
        parser.print_help(sys.stderr)
        exit(1)
    if max_batch < batch_size:
        max_batch = batch_size

    network = args.network
    dev = args.device
    input_time = output_time = 0.
    use_rpc = False
    
    # Check if device is available
    if dev == 'cpu':
        if tvm.cpu(0).exist:
            # dev = Device('cpu', 0)
            ctx = tvm.cpu(0)
            target = 'llvm -mcpu=core-avx2'
        else:
            print('[Error] Device \'%s\' is unrecognizable' % dev)
            exit(1)

    elif dev == 'igpu':
        if tvm.opencl(0).exist:
            # dev = Device('igpu', 0)
            ctx = tvm.opencl(0)
            target = tvm.target.intel_graphics()
        else:
            print('[Error] Device \'%s\' is unrecognizable' % dev)
            exit(1)

    elif dev.find('gpu') >= 0:
        idx_start = dev.find('gpu') + len('gpu')
        gpu_idx = int(dev[idx_start:])
        if tvm.gpu(gpu_idx).exist:
            # dev = Device('gpu', gpu_idx)
            ctx = tvm.gpu(gpu_idx)
            target = 'cuda -device=1050ti'
        else:
            print('[Error] Device \'%s\' is unrecognizable' % dev)
            exit(1)

    elif dev == 'rpc':
        use_rpc = True
        host = '166.104.144.16'
        port = 4123
        
        remote = rpc.connect(host, port)
        ctx = remote.gpu(0)
        target = 'cuda'
        target_host = 'cuda'

    else:
        print('[Error] Device \'%s\' is unrecognizable' % dev)
        exit(1)

    while batch_size <= max_batch:
        # env = Environment(network, batch_size, [dev], '')
        env = Environment(network, batch_size, [], args.log)
        
        # build graph
        net, params, input_shape, output_shape = \
            get_network(name=env.network, batch_size=env.batch_size)
        data = tvm.nd.array((np.random.uniform(size=input_shape)).astype('float32'))
        
        if use_rpc:
            with relay.build_config(opt_level=env.opt_level):
                graph, lib, params = relay.build(net, target=target, target_host=target_host, params=params)

            rpc_thread = threading.Thread(connectRPC, args=(ctx, remote, graph, lib, params, data))
            rpc_thread.start()
            rpc_thread.join()

        else:
            with relay.build_config(opt_level=env.opt_level):
                # graph, lib, params = relay.build(net, target=dev.target, params=params)
                graph, lib, params = relay.build(net, target=target, params=params)

            # module = graph_runtime.create(graph, lib, dev.ctx)
            module = graph_runtime.create(graph, lib, ctx)
            # module.set_input(**params)

            # input time
            input_time = time.time()
            module.set_input('data', data)
            module.set_input(**params)
            ctx.sync()
            input_time = (time.time() - input_time) * 1000 # ms

            # output time
            output_time = time.time()
            module.get_output(0)
            ctx.sync()
            output_time = (time.time() - output_time) * 1000 # ms

        print('input : %7.3f ms\noutput: %7.3f ms\n' % (input_time, output_time))

        if env.log_path != '':
            if path.exists(env.log_path):
                book = openpyxl.load_workbook(env.log_path)
                if env.network in book:
                    sheet = book[env.network]
                else: sheet = book.create_sheet(env.network)
            else:
                book = openpyxl.Workbook()
                sheet = book.create_sheet(env.network)

            row = str(int(env.batch_size/10))
            sheet[str(chr(65)) + row] = batch_size
            sheet[str(chr(65 + 1)) + row] = input_time
            sheet[str(chr(65 + 2)) + row] = output_time

            book.save(env.log_path)

        if increment > 0: batch_size += increment
        else: break
