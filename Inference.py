'''
--- needed updates (20.7.19) ---
  1) Dynamic device allocation
  2) Use threads to do testing more faster
    - if CPU is not used, then sequential builds are no matter
    - else, execute with CPU then builds need to be scheduled
'''

import tvm
from tvm import relay
from tvm.contrib import graph_runtime
from util import get_network
import argparse
import copy
from math import floor, ceil, log2
import numpy as np
import openpyxl
from os import path,_exit
import sys
import threading
import time
from Partition import Partitioner, PerfInfo

class Environment:
    def __init__(self, network, batch_size, devices, log_path=''):
        self.devices = copy.deepcopy(devices)
        self.batch_size = batch_size
        self.network = network
        self.opt_level = 3
        self.test_times = 1
        self.run_times = 1
        self.log_path = log_path

    def getBatches(self):
        batches = []
        for dev in self.devices:
            batches.append(dev.batch_size)
        return batches

    def getMaxTime(self):
        times = []
        for dev in self.devices:
            times.append(dev.result_time.getTime())
        return max(times)

class Device:
    def __init__(self, dev_type='', idx=0):
        self.dev_type = dev_type
        self.idx = idx
        self.batch_size = 0
        self.result_time = None
        self.predict_time = None

        self.eval_time = 0.0
        self.trial = 0
        self.diff = 0.0

        self.getDevInfo()

    def getDevInfo(self):
        if self.dev_type == 'cpu':
            self.ctx = tvm.cpu(self.idx)
            # self.target = 'llvm'
            self.target = 'llvm -mcpu=core-avx2'
            dev_name = self.ctx.device_name
            if dev_name is None:
                dev_name = 'Intel(R) Core(TM) i7-9700K CPU @3.60GHz'
                # dev_name = 'Intel(R) Core(TM) i7-8700K CPU @3.60GHz'
        
        elif self.dev_type == 'igpu':
            self.ctx = tvm.opencl(self.idx)
            self.target = tvm.target.intel_graphics()
            # self.target = tvm.target.intel_graphics(model='Intel® Processor Graphics Gen9')
            dev_name = self.ctx.device_name

        elif self.dev_type == 'gpu':
            cudaInit = tvm.get_global_func('cudaInit')
            cudaInit(self.idx)
            self.ctx = tvm.gpu(self.idx)
            # self.target = 'cuda'
            self.target = 'cuda -device=1050ti'
            dev_name = self.ctx.device_name

        else:
            print('[Error] Unknown Device Type %s' % (self.dev_type))
            exit(1)

        self.name = '[%s] '%(self.dev_type.upper()) + dev_name

    def pushResult(self):
        global result
        string = '%4s %d = %7.2f ms (%6.2f ms)' % (self.dev_type.upper(), self.idx, \
            self.result_time.exec_time + self.result_time.io_time, self.result_time.io_time)
        if self.predict_time > 0:
            string += ' | %7.2f ms' % (self.predict_time)
        string += ' [%3d]\n' % (self.batch_size)
        result += string

    def run(self, graph, lib, params, input_shape, repeat_time=1, mode=''):
        global env, result
        if self.batch_size == 0:
            result += '%s %d = 0 batch\n' % (self.dev_type, self.idx)
            return 0.0

        try: module = graph_runtime.create(graph, lib, self.ctx)
        except Exception as e:
            dev_type = self.dev_type.upper()
            print('\n[Error] Executing with %s (%s %d) failed'
                % (self.name.replace('[%s] '%(dev_type), ''), dev_type, self.idx))
            if mode == 'test': return -1, -1
            else: _exit(1)

        data = tvm.nd.array((np.random.uniform(size=input_shape)).astype('float32'))

        # measure input time
        input_time = time.time()
        module.set_input('data', data)
        self.ctx.sync()
        input_time = time.time() - input_time
        module.set_input(**params)

        timer = module.module.time_evaluator('run', self.ctx, number=1, repeat=repeat_time)
        exec_time = np.mean(np.array(timer().results)) * 1000

        # measure output time
        output_time = time.time()
        module.get_output(0)
        self.ctx.sync()
        output_time = time.time() - output_time

        io_time = (input_time + output_time) * 1000
        self.result_time = PerfInfo(exec_time, io_time)

        if mode != 'test': self.pushResult()
        return exec_time, io_time

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--network', type=str, choices=
                        ['mobilenet', 'squeezenet_v1.0', 'squeezenet_v1.1', 'resnet-18', 'resnet-34',
                        'resnet-50', 'inception_v3', 'vgg-16', 'vgg-19', 'densenet-121'], 
                        help='Name of neural network model to use')
    parser.add_argument('--batch', type=int, help='Batch size')
    parser.add_argument('--device', type=str, default='gpu0',
                        help='Devices to use, write arguments as \'cpu\', \'igpu\' or \'gpu0\' successively')
    parser.add_argument('--log', type=str, default='', help='File path for logging')
    args = parser.parse_args()

    batch_size = args.batch
    if batch_size == 0:
        parser.print_help(sys.stderr)
        exit(1)

    network = args.network
    arg_devs = list(args.device.split(','))

    devices = []
    gpus = []
    cpu = igpu = None
    cpu_idx = igpu_idx = -1

    # Check available devices
    for dev in arg_devs:
        if dev == 'cpu':
            if tvm.cpu(0).exist:
                cpu = Device('cpu', 0)
            else: print('[Error] Device \'%s\' is unrecognizable' % dev)

        elif dev == 'igpu':
            if tvm.opencl(0).exist:
                igpu = Device('igpu', 0)
            else: print('[Error] Device \'%s\' is unrecognizable' % dev)

        elif dev.find('gpu') >= 0:
            idx_start = dev.find('gpu') + len('gpu')
            gpu_idx = int(dev[idx_start:])
            if tvm.gpu(gpu_idx).exist:
                gpus.append(Device('gpu', gpu_idx))
            else: print('[Error] Device \'%s\' is unrecognizable' % dev)
        
        else: print('[Error] Device \'%s\' is unrecognizable' % dev)

    # Add devices
    if cpu is not None:
        cpu_idx = len(devices)
        devices.append(cpu)
    if igpu is not None:
        igpu_idx = len(devices)
        devices.append(igpu)
    if len(gpus) > 0:
        devices.extend(gpus)

    # Check added devices
    if len(devices) == 0:
        parser.print_help(sys.stderr)
        exit(1)

    # Set environment
    env = Environment(network, batch_size, devices, args.log)

    threads = []
    result = ''
    elapsed_time = time.time()
    div_time = 0.0
    
    # Skip partition for just one device
    if len(env.devices) == 1:
        env.devices[0].batch_size = batch_size

    # Start partition
    else:
        div_time = time.time()
        partitioner = Partitioner(env)
        partitioner.startPartition()
        div_time = time.time() - div_time - partitioner.benchmark_time

    # Schedule build order for effective sequential builds
    if igpu_idx >= 0:
        env.devices.insert(len(env.devices)-1, env.devices.pop(igpu_idx))
    if cpu_idx >= 0:
        env.devices.insert(len(env.devices)-1, env.devices.pop(cpu_idx))

    result = '[%s] batch: %d\n' % (env.network, env.batch_size)
    for dev in env.devices:
        if dev.batch_size == 0: continue

        # build_time = time.time()
        net, params, input_shape, output_shape = \
            get_network(name=env.network, batch_size=dev.batch_size)
        with relay.build_config(opt_level=env.opt_level):
            graph, lib, params = relay.build(net, target=dev.target, params=params)
        # build_time = time.time() - build_time
        # print('<%s> build time: %.3f sec' % (dev.name, build_time))

        # Create threads then do run
        t = threading.Thread(target=dev.run, args=(graph, lib, params, input_shape, env.run_times))
        threads.append(t)
        t.start()
    for t in threads:
        t.join()

    elapsed_time = time.time() - elapsed_time

    # temporary codes for logging
    if env.log_path != '':
        if path.exists(env.log_path):
            book = openpyxl.load_workbook(env.log_path)
            if env.network in book:
                sheet = book[env.network]
            else: sheet = book.create_sheet(env.network)
        else:
            book = openpyxl.Workbook()
            sheet = book.create_sheet(env.network)

        # row = str(int(env.batch_size/2))
        row = str(len(env.devices))
        for idx in range(len(env.devices)):
            dev = env.devices[idx]
            sheet[str(chr(idx + 65)) + row] = dev.batch_size
            sheet[str(chr(idx + 65 + len(env.devices))) + row] = dev.result_time.getSummedTime()

        max_time = env.getMaxTime()
        gpu_time = partitioner.estimateDevTime(env.devices[0], env.batch_size)
        sheet[str(chr(65 + len(env.devices)*2)) + row] = gpu_time
        sheet[str(chr(65 + len(env.devices)*2 + 1)) + row] = gpu_time / max_time * 100
        
        book.save(env.log_path)
    
    print(result)
    print('All elapsed time: %.2f sec' % (elapsed_time))
    print('Partitioning time: %.2f sec' % (div_time))
